"""Backend de hoja de cálculo respaldado por un archivo Excel (.xlsx).

Pensado para exportaciones de sistemas como Facturama, que entregan Excel.
Lee la primera fila como encabezados y, al escribir de vuelta, sólo modifica
las celdas de estatus / último recordatorio, preservando el resto del libro.

Requiere ``openpyxl`` (importado de forma perezosa).
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

from ..models import Invoice, parse_amount
from .base import ColumnMap, SpreadsheetBackend, parse_date


class XlsxSpreadsheet(SpreadsheetBackend):
    def __init__(
        self,
        path: str | Path,
        columns: Optional[ColumnMap] = None,
        worksheet: Optional[str] = None,
        header_row: int = 1,
    ):
        self.path = Path(path)
        self.columns = columns or ColumnMap()
        self.worksheet_name = worksheet
        self.header_row = header_row

        self._wb = None
        self._ws = None
        self._header: List[str] = []
        self._col_index: Dict[str, int] = {}  # nombre -> índice 0-based
        self._dirty = False

    # -- carga -----------------------------------------------------------
    def _load(self):
        if self._wb is not None:
            return
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - depende del entorno
            raise ImportError(
                "El backend de Excel requiere 'openpyxl'. "
                "Instálalo con: pip install openpyxl"
            ) from exc

        self._wb = openpyxl.load_workbook(self.path)
        self._ws = (
            self._wb[self.worksheet_name] if self.worksheet_name else self._wb.active
        )
        header_cells = list(self._ws[self.header_row])
        self._header = [
            (str(c.value).strip() if c.value is not None else "") for c in header_cells
        ]
        self._col_index = {name: i for i, name in enumerate(self._header) if name}

    # -- lectura ---------------------------------------------------------
    def read_invoices(self) -> List[Invoice]:
        self._load()
        c = self.columns

        def cell(row: tuple, header: str):
            idx = self._col_index.get(header)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        invoices: List[Invoice] = []
        # min_row = fila siguiente al encabezado; los datos empiezan ahí.
        for offset, row in enumerate(
            self._ws.iter_rows(min_row=self.header_row + 1, values_only=True)
        ):
            # Saltar filas totalmente vacías.
            if row is None or all(v in (None, "") for v in row):
                continue
            invoices.append(
                Invoice(
                    invoice_id=_s(cell(row, c.invoice_id)),
                    client_name=_s(cell(row, c.client_name)),
                    client_email=_s(cell(row, c.client_email)),
                    amount=parse_amount(cell(row, c.amount)),
                    currency=_s(cell(row, c.currency)) or "MXN",
                    issue_date=parse_date(cell(row, c.issue_date)),
                    due_date=parse_date(cell(row, c.due_date)),
                    status=_s(cell(row, c.status)) or "pendiente",
                    last_reminder_stage=_s(cell(row, c.last_reminder_stage)),
                    last_reminder_date=parse_date(cell(row, c.last_reminder_date)),
                    payment_link=_s(cell(row, c.payment_link)),
                    notes=_s(cell(row, c.notes)),
                    # Fila real en la hoja (1-based) para escribir de vuelta.
                    row_ref=self.header_row + 1 + offset,
                )
            )
        return invoices

    # -- escritura -------------------------------------------------------
    def _ensure_column(self, header: str) -> int:
        """Devuelve el índice 0-based de la columna, creándola si no existe."""
        if header in self._col_index:
            return self._col_index[header]
        new_idx = len(self._header)
        self._header.append(header)
        self._col_index[header] = new_idx
        # Escribir el encabezado nuevo en la hoja.
        self._ws.cell(row=self.header_row, column=new_idx + 1, value=header)
        return new_idx

    def stage_update(
        self,
        invoice: Invoice,
        *,
        status: Optional[str] = None,
        last_reminder_stage: Optional[str] = None,
        last_reminder_date: Optional[date] = None,
    ) -> None:
        if invoice.row_ref is None:
            raise ValueError("La factura no tiene fila asociada (¿se leyó del Excel?)")
        self._load()
        row = int(invoice.row_ref)
        c = self.columns
        updates = {
            c.status: status,
            c.last_reminder_stage: last_reminder_stage,
            c.last_reminder_date: (
                last_reminder_date.strftime("%Y-%m-%d") if last_reminder_date else None
            ),
        }
        for header, value in updates.items():
            if value is None:
                continue
            col_idx = self._ensure_column(header)
            self._ws.cell(row=row, column=col_idx + 1, value=value)
        self._dirty = True

    def commit(self) -> None:
        if not self._dirty or self._wb is None:
            return
        self._wb.save(self.path)
        self._dirty = False


def _s(value) -> str:
    """Normaliza una celda a texto sin espacios (fechas -> ISO)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()
